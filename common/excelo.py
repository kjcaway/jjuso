from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED, BLUE
from datetime import datetime

class Excelo:
    wb = None
    ws = None
    isNew = False

    def __init__(self, filePath=None):
        if filePath is None:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.isNew = True
        else:
            self.wb = load_workbook(filePath, data_only=True)
            self.isNew = False

    def setSheet(self, sheetName):
        self.ws = self.wb[sheetName]

    def getSheetList(self):
        return self.wb.get_sheet_names()

    def getMaxColumn(self):
        return self.ws.max_column

    def getAddressList(self, column):
        return self.ws[column]

    def save(self):
        if self.isNew:
            now = datetime.now().strftime('%Y%m%d%H%M%S')
            self.wb.save("new_"+ str(now) + ".xlsx")

    def setExcelShape(self):
        self.ws['A1'].font = Font(color=BLUE)
        self.ws['B1'].font = Font(color=BLUE)
        self.ws['C1'].font = Font(color=BLUE)
        self.ws['D1'].font = Font(color=RED)
        self.ws.column_dimensions['A'].width = 40
        self.ws.column_dimensions['B'].width = 52
        self.ws.column_dimensions['C'].width = 40
        self.ws.column_dimensions['D'].width = 80